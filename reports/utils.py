import os
from io import BytesIO

from django.core.mail import EmailMessage
from django.utils import timezone
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.workbook import Workbook

from reports.models import Exceptions, Vehicle

URL = "https://my.geotab.com/apiv1"

EXCEPTIONS_BODY = {
    "method": "Get",
    "params": {
        "typeName": "Trip",
        "credentials": {
            "database": "moove",
            "sessionId": "2nR_L-I6A8F0K5DVF8srFQ",
            "userName": "moovechallengeuser@mooveconnected.com"
        },
        # TODO : Make an auto 30 days check.
        "search": {
            "fromDate": "2022-08-14T22:00:00.000Z",
            "toDate": "2022-08-22T22:00:00.000Z"
        }
    }
}

TRIPS_BODY = {
    "method": "Get",
    "params": {
        "typeName": "ExceptionEvent",
        "credentials": {
            "database": "moove",
            "sessionId": "2nR_L-I6A8F0K5DVF8srFQ",
            "userName": "moovechallengeuser@mooveconnected.com"
        },
        # TODO : Make an auto 30 days check.
        "search": {
            "fromDate": "2022-08-14T22:00:00.000Z",
            "toDate": "2022-08-22T22:00:00.000Z"
        }
    }
}

VEHICLE_BODY = {
    "method": "Get",
    "params": {
        "typeName": "Device",
        "credentials": {
            "database": "moove",
            "sessionId": "2nR_L-I6A8F0K5DVF8srFQ",
            "userName": "moovechallengeuser@mooveconnected.com"
        }
    }
}

HARSH_ACCELERATION = 'apUro_0nXOUmLV4SVlzK8Xw'
SPEEDING = 'abHSbCv2PKUWKSSGJMoiBnQ'


def generate_excel_file(email, trips):
    """Generate excel file."""
    excelfile = BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    row_num = 1
    worksheet = workbook.create_sheet(title=email, index=1)
    workbook.security.lockStructure = False
    worksheet.protection.sheet = False
    worksheet.protection.formatCells = False

    worksheet.sheet_properties.tabColor = '1072BA'
    worksheet.freeze_panes = 'I2'

    columns = ['License plate', 'Trip Start Date Time', 'Distance Driven(km)',
               'Driving Exception(Speeding) Counts',
               'Driving Exception(HarshAcceleration) Counts']

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.font = Font(bold=True)

    # Iterate through all trips.
    for _, trip in enumerate(trips, 1):
        row_num += 1
        vehicle = Vehicle.objects.filter(
            geo_tab_id=trip.device.geo_tab_id).first()

        # Define the data for each cell in the row
        row = [
            vehicle.license_plate,
            str(trip.start),
            str(trip.stop),
            trip.distance,
            Exceptions.objects.filter(
                device__geo_tab_id=trip.device.geo_tab_id,
                rule___id=SPEEDING).count(),
            Exceptions.objects.filter(
                device__geo_tab_id=trip.device.geo_tab_id,
                rule___id=HARSH_ACCELERATION).count()
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(excelfile)
    now = timezone.now()

    # Send an email.
    message = EmailMessage(
        f'Generate report as of {now.date().isoformat()}',
        f'Generated at: {now.isoformat()}',
        os.getenv('DEFAULT_FROM_EMAIL'),
        [email],
    )
    message.attach(f'{email}.xlsx', excelfile.getvalue(),
                   'application/vnd.ms-excel')
    message.send()
